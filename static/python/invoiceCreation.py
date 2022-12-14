from openpyxl import load_workbook
import shutil
import datetime as date

class createInvoice:
    invoiceTemplatePath = "././invoice/template"
    invoiceOutputPath = "././invoice/output"
    monthArray = ["January", "February", "March", "April", "May", "June", "July", "August", "September", "October", "November", "December"]
    pcs = 0
    cs = 0
    fz = 0
    nv = 0
    jr = 0
    pep = 0
    ad = 0
    pw = 0
    st = 0
    cir = 0
    rpt = 0
    inter = 0
    year = 0

    def __init__(self, season: str, month: str, name: str, rate: str, comments: str, sessions: dict):
        self.season = season
        self.month = month
        self.name = name
        self.rate = rate
        self.comments = comments
        self.sessions = sessions
        
        #get number of days
        self.numberDays = len(list(self.sessions))
        self.days = list(self.sessions)

    """
    Returns invoice output path (string)
    """
    def getInvoiceOutputPath(self):
        return self.invoiceOutputPath

    """
    Gets year in the season based on month
    """
    def getYear(self):
        month = self.monthArray.index(self.month)+1
        if (month >= 9):
            self.year = self.season.split('-')[0]
        else:
            self.year = self.season.split('-')[1]
    
    """
    Copies relevent template and puts it to the output folder
    """
    def openTemplate(self):
        self.getYear()
        self.invoiceOutputPath = self.invoiceOutputPath+"/"+f"{self.month} {self.year} Invoice.xlsx"
        shutil.copyfile(self.invoiceTemplatePath+"/"+f"{self.season}.xlsx", self.invoiceOutputPath)

    """
    Gets total number of sessions of all possible session values set by user
    """
    def setValues(self):
        
        for y in range(self.numberDays):
            numberSessions = len(self.sessions[list(self.sessions)[y]]["sessions"])
            currDay = self.sessions[self.days[y]]["sessions"]
            for x in range(numberSessions):
                if currDay[f"session{x}"]["type"] == "PCS":
                    self.pcs+=int(currDay[f"session{x}"]["amount"])
                elif currDay[f"session{x}"]["type"] == "CS":
                    self.cs+=int(currDay[f"session{x}"]["amount"])
                elif currDay[f"session{x}"]["type"] == "FZ":
                    self.fz+=int(currDay[f"session{x}"]["amount"])
                elif currDay[f"session{x}"]["type"] == "NV":
                    self.nv+=int(currDay[f"session{x}"]["amount"])
                elif currDay[f"session{x}"]["type"] == "JR":
                    self.jr+=int(currDay[f"session{x}"]["amount"])
                elif currDay[f"session{x}"]["type"] == "PEP":
                    self.pep+=int(currDay[f"session{x}"]["amount"])
                elif currDay[f"session{x}"]["type"] == "AD":
                    self.ad+=int(currDay[f"session{x}"]["amount"])
                elif currDay[f"session{x}"]["type"] == "PW":
                    self.pw+=int(currDay[f"session{x}"]["amount"])
                elif currDay[f"session{x}"]["type"] == "ST":
                    self.st+=int(currDay[f"session{x}"]["amount"])
                elif currDay[f"session{x}"]["type"] == "CIR":
                    self.cir+=int(currDay[f"session{x}"]["amount"])
                elif currDay[f"session{x}"]["type"] == "RPT":
                    self.rpt+=int(currDay[f"session{x}"]["amount"])
                elif currDay[f"session{x}"]["type"] == "IN":
                    self.inter+=int(currDay[f"session{x}"]["amount"])
    
    """
    Fills in the information on the Timesheet spreadsheet
    """   
    def fillInvoice(self):
        #to create logic in setting values from site to excel (must check season and have correct template)
        newInvoice = load_workbook(filename = self.invoiceOutputPath)
        newInvoice.active = newInvoice["Timesheet"]
        wb = newInvoice.active
        wb["J5"] = self.name
        wb["J7"] = self.rate
        if self.season == "2022-2023":
            #fill invoice based on 2022-2023 template
            self.setValues()
            if self.pcs > 0:
                wb["C23"] = self.pcs
            if self.cs > 0:
                wb["C24"] = self.cs
            if self.fz > 0:
                wb["C25"] = self.fz
            if self.nv > 0:
                wb["C26"] = self.nv
            if self.jr >0:
                wb["C27"] = self.jr
            if self.pep >0:
                wb["C28"] = self.pep
            if self.ad >0:
                wb["C29"] = self.ad
            if self.pw >0:
                wb["C30"] = self.pw
            if self.st >0:
                wb["C31"] = self.st
            if self.cir >0:
                wb["C32"] = self.cir
            if self.rpt >0:
                wb["C33"] = self.rpt 
        else:
            #fill invoice based on 2019-2020 template for years 2019-2022
            self.setValues()
            if self.pcs > 0:
                wb["C20"] = self.pcs
            if self.cs > 0:
                wb["C21"] = self.cs
            if self.inter > 0:
                wb["C22"] = self.inter
            if self.pep >0:
                wb["C23"] = self.pep
            if self.ad >0:
                wb["C24"] = self.ad
            if self.pw >0:
                wb["C25"] = self.pw
            if self.st >0:
                wb["C26"] = self.st

        # fix comments later
        commentArray = []
        #42 is max characters that can be displayed at once in comment line
        if (len(self.comments) > 42):
            for i in range(0,len(self.comments), 42):
                commentArray.append(self.comments[i:i+42])
            if self.season == "2022-2023":
                for x in range(len(commentArray)):
                    if x <= 6:
                        wb[f"I2{str(3+x)}"] = commentArray[x]
            else:
                for x in range(len(commentArray)):
                    if x <= 4:
                        wb[f"I2{str(x)}"] = commentArray[x]
        else:
            if self.season == "2022-2023":
                wb["I23"] = self.comments
            else:
                wb["I20"] = self.comments

        newInvoice.save(filename = self.invoiceOutputPath)
        self.fillCalendar()
        newInvoice.close()

    """
    Fills the Calendar on the Calendar Spreadsheet
    """
    def fillCalendar(self):
        #to create logic in creating the calendar in the other spreadsheet and to fill in correct and respective worked session dates
        row = ["9", "11", "13", "15", "17"]
        col = ["B", "C", "D", "E", "F", "G", "H"]
        newInvoice = load_workbook(filename = self.invoiceOutputPath)
        newInvoice.active = newInvoice["Calendar"]
        wb = newInvoice.active

        wb["E7"] = self.month

        #get first day of the month
        firstDay = date.datetime(int(self.year), self.monthArray.index(self.month)+1, 1).weekday()
        
        colIndex = self.setFirstDay(firstDay,wb)

        sessionText = ""
        firstSessionDict = self.getSessionOnDay(1)
        #check if there are any sessions done on the first day
        if (firstSessionDict):
            for y in firstSessionDict:
                sessionText += f"{y} - {firstSessionDict[y]}\n"
                #write in cell below day number
            if colIndex == 1:
                wb["C10"] = sessionText
            elif colIndex == 2:
                wb["D10"] = sessionText
            elif colIndex == 3:
                wb["E10"] = sessionText
            elif colIndex == 4:
                wb["F10"] = sessionText
            elif colIndex == 5:
                wb["G10"] = sessionText
            elif colIndex == 6:
                wb["H10"] = sessionText
            elif colIndex == 0:
                wb["B10"] = sessionText

        #if first day is not on a saturday, set next day to be next element in list. otherwise, go back to the first element
        if (colIndex != 6):
            colIndex +=1
            rowIndex = 0
        else:
            colIndex = 0
            rowIndex = 1


        lastDay = self.getLastDay()

        for x in range(2,lastDay+1):
            sessionDict = {}
            sessionText = ""
            if rowIndex <= 4:
                wb[col[colIndex]+row[rowIndex]] = x
                #add comment under day if there is session
                sessionDict = self.getSessionOnDay(x)
                #check if there are any sessions done on that day
                if (sessionDict):
                    for y in sessionDict:
                        if y != "Covering":
                            sessionText += f"{y} - {sessionDict[y]}\n"
                        else:
                            if sessionDict[y] != "":
                                sessionText += f"{y} - {sessionDict[y]}\n"
                        #write in cell below day number
                    wb[col[colIndex]+str(int(row[rowIndex])+1)] = sessionText
                else:
                    wb[col[colIndex]+str(int(row[rowIndex])+1)] = ""
            #if overflow
            if rowIndex > 4:
                rowIndex = 5
                wb[col[colIndex]+row[rowIndex-1]] = str(wb[col[colIndex]+row[rowIndex-1]].value)+"/"+str(x)
                #add comment under day if there is session
                sessionDict = self.getSessionOnDay(x)
                #check if there are any sessions done on that day
                if sessionDict:
                    for y in sessionDict:
                        sessionText += f"{y} - {sessionDict[y]}\n"
                        #write in cell below day number
                    wb[col[colIndex]+str(int(row[rowIndex-1])+1)] = str(wb[col[colIndex]+str(int(row[rowIndex-1])+1)].value) + "\n" + sessionText
                
            colIndex +=1
            if colIndex > 6:
                colIndex = 0
                rowIndex += 1

        newInvoice.save(filename = self.invoiceOutputPath)
        newInvoice.close()

    """
    Sets the first day of the month on the Calendar spreadsheet and returns the column index.
    Arguments:
        firstDay (int) - first day of the month
        wb (openpyxl object) - calendar spreadsheet
    """
    def setFirstDay(self, firstDay: int, wb):
        if firstDay == 0:
            wb["C9"] = 1
            return 1
        elif firstDay == 1:
            wb["D9"] = 1
            return 2
        elif firstDay == 2:
            wb["E9"] = 1
            return 3
        elif firstDay == 3:
            wb["F9"] = 1
            return 4
        elif firstDay == 4:
            wb["G9"] = 1
            return 5
        elif firstDay == 5:
            wb["H9"] = 1
            return 6
        elif firstDay == 6:
            wb["B9"] = 1
            return 0

    """
    Gets the first day column in the calendar spreadsheet
    Arguments:
        firstday (int) - first day of the month
    Returns the column where the first day is on the spreadsheet
    """
    def getFirstDay(self, firstDay):
        if firstDay == 0:
            return "C9"
        elif firstDay == 1:
            return "D9"
        elif firstDay == 2:
            return "E9"
        elif firstDay == 3:
            return "F9"
        elif firstDay == 4:
            return "G9"
        elif firstDay == 5:
            return "H9"
        elif firstDay == 6:
            return "B9"

    """
    Gets the last day of the month
    """
    def getLastDay(self):
        if (self.monthArray.index(self.month)+2 > 12):
            return (date.date(int(self.year), 1, 1) - date.timedelta(days=1)).day
        else:
            return (date.date(int(self.year), self.monthArray.index(self.month)+2, 1) - date.timedelta(days=1)).day
    
    """
    Returns month number from given month
    """
    def getMonthNum(self):
        if len(str(self.monthArray.index(self.month)+1)) == 1:
            monthNum = "0" + str(self.monthArray.index(self.month)+1)
        else:
            monthNum = str(self.monthArray.index(self.month)+1)
        
        return monthNum

    """
    Gets what session has been done on the day given.
    Arguments:
        day (int) - day of the month
    """
    def getSessionOnDay(self, day: int):
        exportDict = {}
        if len(str(day)) == 1:
            #lol
            day = str("0"+str(day))
        
        month = self.getMonthNum()
        
        getDayDict = {}
        
        if f"{self.year}-{month}-{day}" in self.sessions:
            getDayDict = self.sessions[f"{self.year}-{month}-{day}"]["sessions"]
            for x in getDayDict:
                #check if session is already in exportDict
                if getDayDict[x]['type'] in exportDict:
                    exportDict[getDayDict[x]['type']] += getDayDict[x]['amount']
                else:
                    exportDict.update({getDayDict[x]['type']:getDayDict[x]['amount']})
            
            #add name of cover being covered (will return set value "" to cover key if not covering anyone)
            exportDict['Covering'] = self.sessions[f"{self.year}-{month}-{day}"]["cover"]
        return(exportDict)