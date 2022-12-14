from static.python.invoiceCreation import createInvoice
import static.python.invoiceSummary as summary
from openpyxl import load_workbook
import shutil
import datetime as date

class invoiceUpload(createInvoice):
    filename = ""
    uploadPath = "invoice/upload/"

    def __init__(self,  filename: str, uploadedFile = True, season: str= "", month: str= "", name: str= "", rate: str= "", comments: str= "", sessions: dict={}):
        self.filename = filename
        self.uploadPath += filename
        if (not uploadedFile):
            self.uploadPath = self.uploadPath.replace("upload", "output")
        super().__init__(season, month, name, rate, comments, sessions)
    
    def getUploadPath(self):
        return self.uploadPath

    def getSeason(self):
        tempName = self.filename.replace("_", " ")

        self.month = tempName.split(" ")[0]
        self.year = tempName.split(" ")[1]

        index = self.monthArray.index(self.month) +1

        if (index >= 9):
            self.season = f"{self.year}-{int(self.year)+1}"
        else:
            self.season = f"{int(self.year)-1}-{self.year}"
        
        return self.season

    def openUploadedFile(self):
        try:
            invoice = load_workbook(filename = self.uploadPath)
            invoice.active = invoice["Timesheet"]
            wb = invoice.active
            self.name = wb["J5"].value
            self.rate = wb["J7"].value
            if (self.season == "2022-2023"):
                #season is 2022-2023
                self.pcs = wb["C23"].value
                self.cs = wb["C24"].value
                self.fz = wb["C25"].value
                self.nv = wb["C26"].value
                self.jr = wb["C27"].value
                self.pep = wb["C28"].value
                self.ad = wb["C29"].value
                self.pw = wb["C30"].value
                self.st = wb["C31"].value
                self.cir = wb["C32"].value
                self.rpt = wb["C33"].value
            else:
                self.pcs = wb["C20"].value
                self.cs = wb["C21"].value
                self.inter = wb["C22"].value
                self.pep = wb["C23"].value
                self.ad = wb["C24"].value
                self.pw = wb["C25"].value
                self.st = wb["C26"].value
            
            if self.pcs == None:
                self.pcs = 0
            if self.cs == None:
                self.cs = 0
            if self.fz == None:
                self.fz = 0
            if self.nv == None:
                self.nv = 0
            if self.jr == None:
                self.jr = 0
            if self.pep == None:
                self.pep = 0
            if self.ad == None:
                self.ad = 0
            if self.pw == None:
                self.pw = 0
            if self.st == None:
                self.st = 0
            if self.cir == None:
                self.cir = 0
            if self.rpt == None:
                self.rpt = 0
            if self.inter == None:
                self.inter = 0

            #print(f"{self.name} {self.rate} {self.pcs} {self.cs} {self.fz} {self.nv} {self.jr} {self.pep} {self.ad} {self.pw} {self.st} {self.cir} {self.rpt} {self.inter}")
            invoice.close()
        except FileNotFoundError:
            print("Can't find file")
        except Exception as e:
            print (e)

    def getUploadedInvoiceInfo(self):
        if self.season == "2022-2023":
            return {"name": self.name, "month": self.month, "year": self.year, "rate": self.rate, "pcs": self.pcs, "cs": self.cs,
            "fz": self.fz, "nv": self.nv, "jr": self.jr, "pep": self.pep, "ad": self.ad, "pw": self.pw, "st": self.st, "cir": self.cir, "rpt": self.rpt}
        else:
             return {"name": self.name, "month": self.month, "year": self.year, "rate": self.rate, "pcs": self.pcs, "cs": self.cs,
            "pep": self.pep, "ad": self.ad, "pw": self.pw, "st": self.st, "inter": self.inter}

    def updateDatabase(self):
        updateSummary = summary.summaryInvoice()
        updateSummary.insertNewInvoice(self.season, self.month, self.uploadPath)
        updateSummary.createSeasonTable(self.season)
        updateSummary.fillSeasonTable(self.season, self.month, self.sessions)
        del updateSummary

    def getCalendar(self):
        sessionDict = {}
        colArray = ["B9", "C9", "D9", "E9", "F9", "G9", "H9"]
        #get first day
        firstDay = date.datetime(int(self.year), self.monthArray.index(self.month)+1, 1).weekday()
        lastDay = self.getLastDay()

        #getting starting position in calendar
        colIndex = colArray.index(self.getFirstDay(firstDay))

        invoice = load_workbook(filename = self.uploadPath)
        invoice.active = invoice["Calendar"]
        wb = invoice.active
        row = ["9", "11", "13", "15", "17"]
        col = ["B", "C", "D", "E", "F", "G", "H"]

        month = self.getMonthNum()

        #get first day session info
        if not wb[col[colIndex]+str(int(row[0])+1)].value == None:
            sessionDict.update({f"{self.year}-{month}-01":{"sessions": {}, "cover": ""}})
            #LOL
            numSession = 0
            for w in wb[col[colIndex]+str(int(row[0])+1)].value.splitlines():
                if "Covering" in w:
                    sessionDict[f"{self.year}-{month}-01"]["cover"] = w.replace("\n", '').replace("Covering - ", "")
                else:
                    sessionDict[f"{self.year}-{month}-01"]["sessions"].update({f"session{numSession}":{}})
                    sessionDict[f"{self.year}-{month}-01"]["sessions"][f"session{numSession}"].update({"type": w.split(" - ")[0].replace("\n",''), "amount": w.split(" - ")[1].replace("\n",'')})
                    numSession+=1

        #if first day is not on a saturday, set next day to be next element in list. otherwise, go back to the first element
        if (colIndex != 6):
            colIndex +=1
            rowIndex = 0
        else:
            colIndex = 0
            rowIndex = 1
        
        for x in range(1,lastDay+1):
            if rowIndex <= 4:
                #If there are sessions on that day, get those sessions
                if not wb[col[colIndex]+str(int(row[rowIndex])+1)].value == None:
                    #check if the day contains a '/' (has overflow)
                    if "/" in str(wb[col[colIndex]+row[rowIndex]].value):
                        #check if there are sessions for the day before the overflow
                        if (wb[col[colIndex]+str(int(row[rowIndex])+1)].value.split("\n")[0] != ""):
                            #get the sessions for the day before the overflow (values before the '' part of the splitstring)
                            sessionsOverFlowDay = wb[col[colIndex]+str(int(row[rowIndex])+1)].value.splitlines()
                            day = f"{self.year}-{month}-{str(wb[col[colIndex]+row[rowIndex]].value.split('/')[0])}"
                            sessionDict.update({day:{"sessions": {}, "cover": ""}})
                            numSessionsInDay = 0
                            for w in sessionsOverFlowDay:
                                if w == "":
                                    #after the '' part of the splitstring are sessions for the day (after the overflow)
                                    break
                                elif w != "":
                                    #check if there's coverage
                                    if ("Covering" in w):
                                        sessionDict[day]["cover"] = w.replace("\n", '').replace("Covering - ", "")
                                    else:
                                        sessionDict[day]["sessions"].update({f"session{numSessionsInDay}":{}})
                                        sessionDict[day]["sessions"][f"session{numSessionsInDay}"].update({"type": w.split(" - ")[0].replace("\n",''), "amount": w.split(" - ")[1].replace("\n",'')})
                                        numSessionsInDay+=1
                    else:
                        #set key name to the date and value to the sessions on that day
                        if len(str(wb[col[colIndex]+row[rowIndex]].value)) == 1:
                            day = "0"+str(wb[col[colIndex]+row[rowIndex]].value)
                        else:
                            day = str(wb[col[colIndex]+row[rowIndex]].value)

                        sessionDict.update({f"{self.year}-{month}-{day}":{"sessions": {}, "cover": ""}})

                        sessionsInDay = wb[col[colIndex]+str(int(row[rowIndex])+1)].value.splitlines()

                        numSessionsInDay = 0
                        for w in sessionsInDay:
                            if ("Covering" in w):
                                sessionDict[f"{self.year}-{month}-{day}"]["cover"] = w.replace("\n", '').replace("Covering - ","")
                            else:
                                sessionDict[f"{self.year}-{month}-{day}"]["sessions"].update({f"session{numSessionsInDay}":{}})
                                sessionDict[f"{self.year}-{month}-{day}"]["sessions"][f"session{numSessionsInDay}"].update({"type": w.split(" - ")[0].replace("\n",''), "amount": w.split(" - ")[1].replace("\n",'')})
                                numSessionsInDay+=1

            #if overflow
            if rowIndex > 4:
                rowIndex = 5
                
                #If there are sessions on that day, get those sessions
                if not wb[col[colIndex]+str(int(row[rowIndex-1])+1)].value == None:
                    #check if there are sessions for the day after the overflow
                    if (wb[col[colIndex]+str(int(row[rowIndex-1])+1)].value.split("\n")[len(wb[col[colIndex]+str(int(row[rowIndex-1])+1)].value.split("\n"))-1] != ""):
                        sessionsOverFlowDay = wb[col[colIndex]+str(int(row[rowIndex-1])+1)].value.splitlines()
                        atSessionsOverFlowText = False
                        y=0
                        day = f"{self.year}-{month}-{str(wb[col[colIndex]+row[rowIndex-1]].value.split('/')[1])}"
                        sessionDict.update({day:{"sessions": {}, "cover": ""}})
                        for w in sessionsOverFlowDay:
                            if atSessionsOverFlowText:
                                if ("Covering" in w):
                                        sessionDict[day]["cover"] = w.replace("\n", '').replace("Covering - ", "")
                                else:
                                    sessionDict[day]["sessions"].update({f"session{y}":{}})
                                    sessionDict[day]["sessions"][f"session{y}"].update({"type": w.split(" - ")[0].replace("\n",''), "amount": w.split(" - ")[1].replace("\n",'')})
                                    y+=1

                            elif w == "":
                                atSessionsOverFlowText = True
            colIndex +=1
            if colIndex > 6:
                colIndex = 0
                rowIndex += 1
        
        invoice.close()
        self.sessions = sessionDict
        return(sessionDict)
